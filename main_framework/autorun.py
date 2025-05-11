import os
import json
import subprocess
from openpyxl import load_workbook

def run_experiments(file_path, start_image_id, end_image_id):
    # 加载 Excel 文件
    wb = load_workbook(file_path)
    sheet = wb.active

    # 找到 ID 和结果列的索引
    id_column = None
    determination_breach_column = None
    suggestion_breach_column = None
    for col_idx, cell in enumerate(sheet[1], start=1):
        if cell.value == "ID":
            id_column = col_idx
        if cell.value == "Model Results Interpretation_Breach":
            determination_breach_column = col_idx
        if cell.value == "ModelSuggestion_Breach":
            suggestion_breach_column = col_idx

    if id_column is None or determination_breach_column is None or suggestion_breach_column is None:
        print("Error: Required columns (ID, Model Results Interpretation_Breach, ModelSuggestion_Breach) are missing.")
        return

    # 找到起始和终止 ID 的行号
    start_row = None
    end_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row[id_column - 1].value == start_image_id:
            start_row = row_idx
        if row[id_column - 1].value == end_image_id:
            end_row = row_idx
        if start_row and end_row:
            break

    if start_row is None:
        print(f"Error: Start image_id '{start_image_id}' not found in the ID column.")
        return

    if end_row is None:
        print(f"Error: End image_id '{end_image_id}' not found in the ID column.")
        return

    # 逐行运行实验
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=id_column, max_col=id_column):
        image_id = row[0].value
        if not image_id:
            continue

        print(f"Running experiment for image_id: {image_id}")
        try:
            # 执行 Python 脚本
            subprocess.run(["python", "run_all_steps.py", str(image_id)], check=True)

            # 读取 output_result.json 文件
            with open("output/output_result.json", "r") as json_file:
                result_data = json.load(json_file)

            # 确保 result_data 是一个列表，并找到对应 image_id 的结果
            if isinstance(result_data, list):
                matched_result = next((item for item in result_data if item.get("image_ID") == str(image_id)), None)
                if matched_result:
                    determination = matched_result.get("determination", "N/A")
                    suggestion = matched_result.get("suggestion", "N/A")
                else:
                    determination = "N/A"
                    suggestion = "N/A"
            else:
                print(f"Unexpected data format in output_result.json for image_id {image_id}")
                determination = "N/A"
                suggestion = "N/A"

            # 将结果写回 Excel 文件
            determination_cell = sheet.cell(row=row[0].row, column=determination_breach_column)
            suggestion_cell = sheet.cell(row=row[0].row, column=suggestion_breach_column)
            determination_cell.value = determination
            suggestion_cell.value = suggestion

        except subprocess.CalledProcessError as e:
            print(f"Error occurred while running for image_id {image_id}: {e}")
        except FileNotFoundError:
            print("Error: output/output_result.json not found.")
        except json.JSONDecodeError:
            print("Error: Failed to decode JSON from output_result.json.")

    # 保存更新后的 Excel 文件
    wb.save(file_path)
    print("Experiments completed and results saved.")

if __name__ == "__main__":
    # 输入文件路径、起始 ID 和终止 ID
    excel_file_path = "output/experiments.xlsx"
    start_image_id = input("Enter the starting image_id: ")
    end_image_id = input("Enter the ending image_id: ")

    if not os.path.exists(excel_file_path):
        print(f"Error: {excel_file_path} does not exist.")
    else:
        run_experiments(excel_file_path, start_image_id, end_image_id)
